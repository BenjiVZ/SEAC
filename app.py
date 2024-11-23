from flask import Flask, request, render_template, redirect, flash, url_for, send_file, jsonify, session, make_response, Response
import win32com.client
import pythoncom
import os
import tempfile
from werkzeug.utils import secure_filename
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime, time
from datetime import datetime as dt, time as dt_time
import time
import json
import logging
import subprocess
import xlsxwriter
import openpyxl
from io import BytesIO
import base64
import mysql.connector
from mysql.connector import Error
import csv
import io
import calendar
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import requests
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, time as datetime_time

# Importar desde config.py
from config import VBA_CODE, insert_vba_code
from proceso_completo import (
    procesar_archivo_vba,
    procesar_pandas_automatico,
    guardar_df_temporal,
    limpiar_datos_adicionales_automatico,
    convertir_a_csv_automatico,
    tabular_por_dias_automatico,
    configurar_gourmet_automatico,
    limpiar_archivos_temporales
)

app = Flask(__name__)
app.secret_key = os.urandom(24)

# Configurar logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

db_config = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'database': 'seac'
}

@app.route('/')
def index():
    return render_template('index.html')

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'csv'}

url = "https://hebbkx1anhila5yf.public.blob.vercel-storage.com/resultado_tabulado_2024-09-aZrsgHKgTRFqBWKoudNPpJvpulVvlm.csv"
response = requests.get(url)
with open("resultado_tabulado_2024-09.csv", "wb") as f:
    f.write(response.content)

df = pd.read_csv("resultado_tabulado_2024-09.csv")
print(df.dtypes)
print(df['Apertura'].head())
print(df['Cierre'].head())

#========================================================================================================================================================================================================================================================

""" Analizar datos Candelaria Center ⬇️ """

def parse_time(time_str):
    """
    Convierte string de tiempo a objeto datetime.time en formato 24h.
    Maneja múltiples formatos de entrada y siempre retorna en formato 24h.
    
    Args:
        time_str: String con el tiempo en formato HH:MM, HH:MM AM/PM, o HHMM
        
    Returns:
        datetime.time: Objeto time en formato 24h
        None: Si el string es inválido o vacío
    """
    try:
        if pd.isna(time_str) or time_str == '':
            return None
        
        # Limpiar y normalizar el string de tiempo
        time_str = str(time_str).strip().upper()
        
        # Casos posibles:
        # 1. HH:MM (formato 24h)
        # 2. HH:MM AM/PM (formato 12h)
        # 3. HHMM (formato numérico)
        
        # Caso 1: Formato HH:MM (24h)
        if ':' in time_str and len(time_str) <= 5:
            try:
                return datetime.strptime(time_str, '%H:%M').time()
            except ValueError:
                pass
        
        # Caso 2: Formato con AM/PM
        if 'AM' in time_str or 'PM' in time_str:
            try:
                # Remover espacios extra entre hora y AM/PM
                time_str = ' '.join(time_str.split())
                return datetime.strptime(time_str, '%I:%M %p').time()
            except ValueError:
                try:
                    # Intentar sin espacios entre hora y AM/PM
                    return datetime.strptime(time_str, '%I:%M%p').time()
                except ValueError:
                    pass
        
        # Caso 3: Formato numérico HHMM
        if time_str.isdigit() and len(time_str) == 4:
            try:
                hours = int(time_str[:2])
                minutes = int(time_str[2:])
                if 0 <= hours <= 23 and 0 <= minutes <= 59:
                    return datetime_time(hours, minutes)
            except ValueError:
                pass
        
        # Último intento: forzar interpretación como 24h
        try:
            # Remover cualquier carácter no numérico excepto :
            clean_time = ''.join(c for c in time_str if c.isdigit() or c == ':')
            if ':' not in clean_time and len(clean_time) == 4:
                clean_time = f"{clean_time[:2]}:{clean_time[2:]}"
            return datetime.strptime(clean_time, '%H:%M').time()
        except ValueError:
            return None
            
    except Exception as e:
        logger.error(f"Error parsing time '{time_str}': {str(e)}")
        return None

def create_bar_chart(data, title, y_label):
    """Crear gráfico de barras personalizado con Plotly"""
    fig = go.Figure()
    
    # Agregar barras para cada día
    for dia in data['Día'].unique():
        dia_data = data[data['Día'] == dia]
        fig.add_trace(go.Bar(
            name=dia,
            x=dia_data['Nombre'],
            y=dia_data['Total_Incumplimientos'],
            text=dia_data['Total_Incumplimientos'],
            textposition='auto',
        ))
    
    # Personalizar el diseño
    fig.update_layout(
        title=title,
        yaxis_title=y_label,
        xaxis_title="Local",
        barmode='group',
        showlegend=True,
        legend_title="Día",
        height=600,
        template='plotly_white',
        xaxis_tickangle=-45,
        font=dict(size=12)
    )
    
    return fig

def analyze_data(file_path, times, exceptions_list=None):
    """Analiza los datos del archivo CSV y genera resultados con y sin excepciones"""
    if exceptions_list is None:
        exceptions_list = []
    
    # Intentar diferentes codificaciones
    encodings = ['utf-8', 'cp1252', 'latin1', 'iso-8859-1', 'windows-1252']
    df = None
    
    for encoding in encodings:
        try:
            df = pd.read_csv(file_path, encoding=encoding)
            print(f"Archivo leído exitosamente con codificación: {encoding}")
            break
        except UnicodeDecodeError:
            continue
        except Exception as e:
            print(f"Error con codificación {encoding}: {str(e)}")
            continue
    
    if df is None:
        # Si ninguna codificación funcionó, intentar con detección automática
        try:
            import chardet
            with open(file_path, 'rb') as file:
                raw_data = file.read()
                result = chardet.detect(raw_data)
                encoding = result['encoding']
                print(f"Codificación detectada: {encoding}")
                df = pd.read_csv(file_path, encoding=encoding)
        except Exception as e:
            raise Exception(f"No se pudo leer el archivo con ninguna codificación: {str(e)}")

    # Resto de la función igual...
    # Ordenar los días de la semana correctamente para visualización
    day_order = ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SABADO', 'DOMINGO']
    df['Día Semana 1'] = pd.Categorical(df['Día Semana 1'], categories=day_order, ordered=True)
    df = df.sort_values(['Día Semana 1', 'Fecha 1'])
    
    # Convertir columnas de tiempo
    df['Apertura'] = df['Apertura'].apply(parse_time)
    df['Cierre'] = df['Cierre'].apply(parse_time)
    
    results = []
    incumplimientos_por_local = {}
    incumplimientos_con_excepciones = {}
    
    # Mapeo de días para horarios límite
    day_mapping = {
        'Lunes': 'lunes',
        'Martes': 'martes',
        'Miércoles': 'miercoles',
        'Jueves': 'jueves',
        'Viernes': 'viernes',
        'Sábado': 'sabado',
        'Domingo': 'domingo',
        'Feriado': 'feriado',
        'Especial': 'especial'
    }

    for _, row in df.iterrows():
        local = row['N° de Local']
        nombre = row['Nombre']
        dia_original = row['Día'].strip()
        dia = day_mapping.get(dia_original.title())  # Convertir a título para manejar variaciones en mayúsculas
        
        if dia is None:
            print(f"Advertencia: Día no reconocido '{dia_original}' para local {local} - {nombre}")
            continue
            
        fecha = row['Fecha 1']
        apertura = row['Apertura']
        cierre = row['Cierre']
        
        # Inicializar contadores para ambos análisis
        if local not in incumplimientos_por_local:
            incumplimientos_por_local[local] = {
                'local': f"{local} - {nombre}",
                'faltas': 0,
                'faltas_apertura': 0,
                'faltas_cierre': 0
            }
            incumplimientos_con_excepciones[local] = {
                'local': f"{local} - {nombre}",
                'faltas': 0,
                'faltas_apertura': 0,
                'faltas_cierre': 0,
                'excepciones': 0
            }
        
        # Obtener horarios límite
        horario = times.get(dia)
        if not horario:
            print(f"Advertencia: No se encontraron horarios para el día: {dia} (original: {dia_original})")
            continue
        
        hora_limite_apertura = datetime.strptime(horario['apertura'], '%H:%M').time()
        hora_limite_cierre = datetime.strptime(horario['cierre'], '%H:%M').time()
        
        # Resto del código igual...
        excepcion = next((e for e in exceptions_list 
                        if str(e['local']) == str(local) and e['fecha'] == fecha), None)
        
        incumplimiento_apertura = False
        incumplimiento_cierre = False
        no_registrado_apertura = False
        no_registrado_cierre = False
        es_excepcion_apertura = excepcion and excepcion.get('excepcion_apertura', False)
        es_excepcion_cierre = excepcion and excepcion.get('excepcion_cierre', False)
        
        # Verificaciones y conteo igual que antes...
        if apertura is None:
            no_registrado_apertura = True
        elif apertura > hora_limite_apertura:
            incumplimiento_apertura = True
            incumplimientos_por_local[local]['faltas'] += 1
            incumplimientos_por_local[local]['faltas_apertura'] += 1
            
            if not es_excepcion_apertura:
                incumplimientos_con_excepciones[local]['faltas'] += 1
                incumplimientos_con_excepciones[local]['faltas_apertura'] += 1
            else:
                incumplimientos_con_excepciones[local]['excepciones'] += 1

        if cierre is None:
            no_registrado_cierre = True
        elif cierre < hora_limite_cierre:
            incumplimiento_cierre = True
            incumplimientos_por_local[local]['faltas'] += 1
            incumplimientos_por_local[local]['faltas_cierre'] += 1
            
            if not es_excepcion_cierre:
                incumplimientos_con_excepciones[local]['faltas'] += 1
                incumplimientos_con_excepciones[local]['faltas_cierre'] += 1
            else:
                incumplimientos_con_excepciones[local]['excepciones'] += 1

        results.append({
            'N° de Local': local,
            'Nombre': nombre,
            'Día': row['Día'],
            'Tipo_Día': dia,
            'Fecha': fecha,
            'Apertura': str(apertura) if apertura else 'No registrado',
            'Hora_Limite_Apertura': str(hora_limite_apertura),
            'Cierre': str(cierre) if cierre else 'No registrado',
            'Hora_Limite_Cierre': str(hora_limite_cierre),
            'Incumplimiento_Apertura': incumplimiento_apertura,
            'Incumplimiento_Cierre': incumplimiento_cierre,
            'No_Registrado_Apertura': no_registrado_apertura,
            'No_Registrado_Cierre': no_registrado_cierre,
            'Es_Excepcion_Apertura': es_excepcion_apertura,
            'Es_Excepcion_Cierre': es_excepcion_cierre
        })

    # Crear gráficos y resumen
    fig_total, fig_tipo = crear_graficos(incumplimientos_por_local, "Incumplimientos")
    fig_total_exc, fig_tipo_exc = crear_graficos(incumplimientos_con_excepciones, "Incumplimientos (Con Excepciones)")

    resumen_excepciones = [
        {
            'numero': local,
            'nombre': info['local'].split(' - ')[1],
            'total': info['faltas'],
            'apertura': info['faltas_apertura'],
            'cierre': info['faltas_cierre'],
            'excepciones': info['excepciones'],
            'tipos_dias': df[df['N° de Local'] == local]['Día'].unique().tolist()
        }
        for local, info in incumplimientos_con_excepciones.items()
        if info['faltas'] > 0 or info['excepciones'] > 0
    ]

    return {
        'graficos': {
            'incumplimientos_total': fig_total.to_json(),
            'incumplimientos_tipo': fig_tipo.to_json(),
            'incumplimientos_total_excepciones': fig_total_exc.to_json(),
            'incumplimientos_tipo_excepciones': fig_tipo_exc.to_json()
        },
        'tabla': results,
        'tabla_incumplimientos': [
            row for row in results 
            if ((row['Incumplimiento_Apertura'] or row['Incumplimiento_Cierre']) and 
                not (row['Es_Excepcion_Apertura'] or row['Es_Excepcion_Cierre'])) or
            (row['No_Registrado_Apertura'] or row['No_Registrado_Cierre'])
        ],
        'resumen_excepciones': resumen_excepciones
    }

def crear_graficos(datos, titulo_base):
    """Función auxiliar para crear gráficos"""
    locales = []
    faltas = []
    
    for info in datos.values():
        if info['faltas'] > 0:  # Solo mostrar locales con faltas
            locales.append(info['local'])
            faltas.append(info['faltas'])

    # Gráfico de incumplimientos totales
    fig_total = go.Figure(data=[
        go.Bar(
            x=locales,
            y=faltas,
            text=faltas,
            textposition='auto',
        )
    ])
    
    fig_total.update_layout(
        title=f'{titulo_base} por Local',
        xaxis_title='Local',
        yaxis_title='Número de Incumplimientos',
        xaxis_tickangle=-45,
        height=600,
        margin=dict(b=150),
        showlegend=False
    )

    # Gráfico de incumplimientos por tipo
    apertura = []
    cierre = []

    for info in datos.values():
        if info['faltas'] > 0:
            apertura.append(info['faltas_apertura'])
            cierre.append(info['faltas_cierre'])

    fig_tipo = go.Figure(data=[
        go.Bar(name='Apertura', x=locales, y=apertura),
        go.Bar(name='Cierre', x=locales, y=cierre)
    ])

    fig_tipo.update_layout(
        title=f'{titulo_base} por Tipo',
        xaxis_title='Local',
        yaxis_title='Número de Incumplimientos',
        barmode='group',
        xaxis_tickangle=-45,
        height=600,
        margin=dict(b=150)
    )

    return fig_total, fig_tipo

@app.route('/upload')
def upload():
    return render_template('upload.html')

@app.route('/upload_file', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No se seleccionó ningún archivo')
            return render_template('upload.html')
        
        file = request.files['file']
        if file.filename == '':
            flash('No se seleccionó ningún archivo')
            return render_template('upload.html')
        
        if file:
            # Guardar archivo
            file_path = os.path.join('uploads', 'uploaded_file.csv')
            os.makedirs('uploads', exist_ok=True)
            file.save(file_path)
            
            try:
                # Recopilar horarios límite
                times = {}
                days = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado', 'domingo', 'especial', 'feriado']
                
                # Debug: Imprimir los datos del formulario
                print("Datos del formulario:")
                for key in request.form:
                    print(f"{key}: {request.form[key]}")
                
                for day in days:
                    apertura = request.form.get(f'{day}_apertura')
                    cierre = request.form.get(f'{day}_cierre')
                    if apertura and cierre:
                        times[day] = {
                            'apertura': apertura,
                            'cierre': cierre
                        }
                
                # Debug: Imprimir times
                print("Times dictionary:", times)
                
                session['times'] = times
                results = analyze_data(file_path, times)
                
                print(f"Número de incumplimientos: {len(results['tabla_incumplimientos'])}")
                
                return render_template('results.html', 
                                    graficos=results['graficos'],
                                    tabla=results['tabla'],
                                    tabla_incumplimientos=results['tabla_incumplimientos'])
                
            except Exception as e:
                flash(f'Error procesando datos: {str(e)}')
                return render_template('upload.html')
    
    return render_template('upload.html')

@app.route('/agregar_excepcion', methods=['POST'])
def agregar_excepcion():
    try:
        data = request.get_json()
        print("Datos recibidos:", data)  # Para debugging
        
        if not data or 'excepciones' not in data:
            return jsonify({'error': 'No se recibieron datos de excepciones'}), 400
            
        # Crear el directorio si no existe
        os.makedirs('data', exist_ok=True)
        exceptions_file = os.path.join('data', 'exceptions.csv')
        
        # Preparar nuevas excepciones
        new_exceptions = []
        for exc in data['excepciones']:
            if not all(key in exc for key in ['local', 'fecha', 'excepcion_apertura', 'excepcion_cierre']):
                return jsonify({'error': 'Datos de excepción incompletos'}), 400
                
            new_exceptions.append({
                'local': str(exc['local']),  # Convertir a string para asegurar consistencia
                'fecha': exc['fecha'],
                'excepcion_apertura': bool(exc['excepcion_apertura']),  # Convertir a boolean
                'excepcion_cierre': bool(exc['excepcion_cierre']),
                'fecha_registro': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
        
        new_exceptions_df = pd.DataFrame(new_exceptions)
        
        # Leer archivo existente o crear uno nuevo
        if os.path.exists(exceptions_file):
            try:
                exceptions = pd.read_csv(exceptions_file)
            except Exception as e:
                print(f"Error leyendo archivo de excepciones: {e}")
                exceptions = pd.DataFrame(columns=['local', 'fecha', 'excepcion_apertura', 'excepcion_cierre', 'fecha_registro'])
        else:
            exceptions = pd.DataFrame(columns=['local', 'fecha', 'excepcion_apertura', 'excepcion_cierre', 'fecha_registro'])
        
        # Eliminar excepciones existentes para los mismos locales y fechas
        for _, exc in new_exceptions_df.iterrows():
            exceptions = exceptions[~((exceptions['local'].astype(str) == str(exc['local'])) & 
                                   (exceptions['fecha'] == exc['fecha']))]
        
        # Concatenar nuevas excepciones
        exceptions = pd.concat([exceptions, new_exceptions_df], ignore_index=True)
        
        # Guardar excepciones
        try:
            exceptions.to_csv(exceptions_file, index=False)
        except Exception as e:
            print(f"Error guardando excepciones: {e}")
            return jsonify({'error': 'Error al guardar las excepciones'}), 500
        
        # Reanalizar los datos
        file_path = os.path.join('uploads', 'uploaded_file.csv')
        if not os.path.exists(file_path):
            return jsonify({'error': 'No se encontró el archivo de datos'}), 404
        
        times = session.get('times', {})
        if not times:
            return jsonify({'error': 'No se encontraron los horarios configurados'}), 400
        
        # Reanalizar datos con las nuevas excepciones
        results = analyze_data(file_path, times, exceptions.to_dict('records'))
        
        # Devolver los datos actualizados
        return jsonify({
            'success': True,
            'message': 'Excepciones aplicadas correctamente',
            'graficos': {
                'incumplimientos_total_excepciones': results['graficos']['incumplimientos_total_excepciones'],
                'incumplimientos_tipo_excepciones': results['graficos']['incumplimientos_tipo_excepciones']
            },
            'resumen_excepciones': results['resumen_excepciones']
        })

    except Exception as e:
        print(f"Error en agregar_excepcion: {str(e)}")  # Para debugging
        import traceback
        traceback.print_exc()  # Esto imprimirá el stack trace completo
        return jsonify({'error': str(e)}), 500
    
""" Procesar Excel para analizar datos Candelaria Center """

"""     @app.route('/generar_excel_completo', methods=['POST'])
    def generar_excel_completo():
        try:
            datos = request.json
            
            # Crear un nuevo libro de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte Completo"
            
            # Agregar encabezados
            for col, header in enumerate(datos['encabezados'], 1):
                ws.cell(row=1, column=col, value=header)
            
            # Estilos
            from openpyxl.styles import PatternFill, Font
            rojo = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            verde = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
            azul = PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
            blanco = Font(color='FFFFFF')
            
            # Agregar datos y aplicar estilos
            for row_idx, fila in enumerate(datos['filas'], 2):
                for col_idx, (valor, estilo) in enumerate(zip(fila['datos'], fila['estilos']), 1):
                    celda = ws.cell(row=row_idx, column=col_idx, value=valor)
                    if estilo == 'incumplimiento':
                        celda.fill = rojo
                        celda.font = blanco
                    elif estilo == 'excepcion':
                        celda.fill = verde
                        celda.font = blanco
                    elif estilo == 'no-registrado':
                        celda.fill = azul
                        celda.font = blanco
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Guardar en un buffer en memoria
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return send_file(
                excel_buffer,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'reporte_completo_{datetime.now().strftime("%Y%m%d")}.xlsx'
            )
            
        except Exception as e:
            return jsonify({'error': str(e)}), 500
 """

VBA_CODE = """
Sub DesvincularCeldas(sheetName As String)
    Dim ws As Worksheet
    Dim celda As Range

    ' Definir la hoja especificada
    Set ws = ThisWorkbook.Sheets(sheetName)
    
    ' Recorrer cada celda usada en la hoja
    For Each celda In ws.UsedRange
        ' Si la celda contiene una fórmula con vínculo
        If celda.HasFormula Then
            ' Copiar solo el valor de la celda y desvincular la fórmula
            celda.Value = celda.Value
        End If
    Next celda

    MsgBox "Todas las celdas vinculadas han sido desvinculadas y sus valores copiados."
End Sub

Sub TabularInformacionCorregido(sheetName As String)
    Dim ws As Worksheet
    Dim nuevaWs As Worksheet
    Dim ultimaFila As Long
    Dim i As Long, j As Long
    Dim filaNueva As Long
    Dim diaSemana1 As String, diaSemana2 As String
    Dim fecha1 As Date, fecha2 As Date
    Dim apertura As String, cierre As String
    Dim nLocal As String, nombre As String
    
    ' Definir la hoja origen y crear una nueva hoja para la tabla
    Set ws = ThisWorkbook.Sheets(sheetName)
    Set nuevaWs = ThisWorkbook.Sheets.Add
    nuevaWs.Name = "Tabla Tabulada"
    
    ' Escribir encabezados en la nueva hoja
    nuevaWs.Cells(1, 1).Value = "N° de Local"
    nuevaWs.Cells(1, 2).Value = "Nombre"
    nuevaWs.Cells(1, 3).Value = "Apertura"
    nuevaWs.Cells(1, 4).Value = "Cierre"
    nuevaWs.Cells(1, 5).Value = "Día Semana 1"
    nuevaWs.Cells(1, 6).Value = "Día Semana 2"
    nuevaWs.Cells(1, 7).Value = "Fecha 1"
    nuevaWs.Cells(1, 8).Value = "Fecha 2"
    
    filaNueva = 2 ' Empezar a llenar desde la segunda fila
    
    ' Encontrar la última fila con datos en la columna A (número de local)
    ultimaFila = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    
    ' Recorrer las filas de la hoja original (desde la fila 6)
    For i = 6 To ultimaFila
        nLocal = ws.Cells(i, 1).Value
        nombre = ws.Cells(i, 2).Value
        
        ' Recorrer las columnas de días y horarios (C y D en adelante)
        For j = 3 To ws.Cells(4, ws.Columns.Count).End(xlToLeft).Column Step 2
            ' Leer los valores correspondientes
            diaSemana1 = ws.Cells(4, j).Value  ' Día semana apertura
            diaSemana2 = ws.Cells(4, j + 1).Value  ' Día semana cierre
            fecha1 = ws.Cells(5, j).Value  ' Fecha apertura
            fecha2 = ws.Cells(5, j + 1).Value  ' Fecha cierre
            apertura = ws.Cells(i, j).Value  ' Hora apertura
            cierre = ws.Cells(i, j + 1).Value  ' Hora cierre
            
            ' Validar que haya datos de apertura y cierre
            If apertura <> "CERRADO" And apertura <> "" And cierre <> "" Then
                ' Copiar la información a la nueva hoja
                nuevaWs.Cells(filaNueva, 1).Value = nLocal
                nuevaWs.Cells(filaNueva, 2).Value = nombre
                nuevaWs.Cells(filaNueva, 3).Value = apertura
                nuevaWs.Cells(filaNueva, 4).Value = cierre
                nuevaWs.Cells(filaNueva, 5).Value = diaSemana1
                nuevaWs.Cells(filaNueva, 6).Value = diaSemana2
                nuevaWs.Cells(filaNueva, 7).Value = fecha1
                nuevaWs.Cells(filaNueva, 8).Value = fecha2
                filaNueva = filaNueva + 1
            End If
        Next j
    Next i
    
    MsgBox "Información tabulada correctamente en 'Tabla Tabulada'."
End Sub
"""

""" Procesar Excel Codigo nuevo y  ⬇️ """

@app.route('/proceso_completo', methods=['GET', 'POST'])
def proceso_completo():
    if request.method == 'POST':
        step = request.form.get('step', '1')
        
        if step == '1':
            file = request.files.get('file')
            sheet_name = request.form.get('sheet_name')
            
            if not file or not sheet_name:
                return render_template('proceso_paso1.html', 
                    error="Por favor complete todos los campos")
            
            try:
                # Procesar VBA y guardar resultado en sesión
                success, messages, temp_path = procesar_archivo_vba(file, sheet_name)
                if not success:
                    return render_template('proceso_paso1.html', 
                        error=messages[0] if messages else "Error desconocido")
                
                session['temp_file'] = temp_path
                
                # Procesar con pandas
                df = procesar_pandas_automatico(temp_path)
                temp_path_pandas = guardar_df_temporal(df, 'pandas')
                session['temp_file_pandas'] = temp_path_pandas
                
                # Limpiar datos adicionales
                df = limpiar_datos_adicionales_automatico(temp_path_pandas)
                temp_path_limpio = guardar_df_temporal(df, 'limpio')
                session['temp_file_limpio'] = temp_path_limpio
                
                # Convertir a CSV
                temp_path_csv = convertir_a_csv_automatico(temp_path_limpio)
                session['temp_file_csv'] = temp_path_csv
                
                return render_template('proceso_paso2.html', messages=messages)
                
            except Exception as e:
                return render_template('proceso_paso1.html', error=str(e))
                
        elif step == '2':
            # Segundo paso: Configurar mes y feriados
            month = request.form.get('month')
            feriados = request.form.get('feriados', '').split(',')
            
            try:
                temp_path_csv = session.get('temp_file_csv')
                if not temp_path_csv:
                    return render_template('proceso_paso2.html', 
                        error="No se encontró el archivo temporal. Por favor, comience el proceso nuevamente.")
                
                # Tabular por días
                df = tabular_por_dias_automatico(temp_path_csv, month, feriados)
                temp_path_tabulado = guardar_df_temporal(df, 'tabulado')
                session['temp_file_tabulado'] = temp_path_tabulado
                
                # Obtener lista de tiendas para el siguiente paso
                tiendas = df['Nombre'].unique().tolist()
                
                return render_template('proceso_paso3.html', tiendas=tiendas)
                
            except Exception as e:
                return render_template('proceso_paso2.html', error=str(e))
                
        elif step == '3':
            try:
                temp_path_tabulado = session.get('temp_file_tabulado')
                if not temp_path_tabulado:
                    return render_template('proceso_paso3.html', 
                        error="No se encontró el archivo temporal")
                
                tiendas_gourmet = request.form.getlist('tiendas_gourmet')
                df_final = configurar_gourmet_automatico(temp_path_tabulado, tiendas_gourmet)
                
                # Crear buffer en memoria
                output = io.BytesIO()
                
                # Guardar DataFrame como CSV en el buffer
                df_final.to_csv(output, index=False, encoding='utf-8-sig')
                output.seek(0)
                
                # Generar nombre de archivo
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f'resultado_final_{timestamp}.csv'
                
                # Configurar headers específicos
                headers = {
                    'Content-Type': 'text/csv',
                    'Content-Disposition': f'attachment; filename="{filename}"',
                    'Cache-Control': 'no-cache, no-store, must-revalidate',
                    'Pragma': 'no-cache',
                    'Expires': '0'
                }
                
                # Limpiar archivos temporales
                try:
                    limpiar_archivos_temporales(session)
                except Exception as e:
                    logger.error(f"Error limpiando archivos temporales: {str(e)}")
                
                # Retornar la respuesta
                return Response(
                    output.getvalue(),
                    mimetype='text/csv',
                    headers=headers
                )
                
            except Exception as e:
                logger.error(f"Error en paso 3: {str(e)}")
                return render_template('proceso_paso3.html', 
                    error=f"Error al procesar el archivo: {str(e)}", 
                    tiendas=tiendas)
    
    # Si es GET, mostrar el primer paso
    return render_template('proceso_paso1.html')

""" Procesar Excel Codigo antiguo y sin usar ⬇️ """

""" def insert_vba_code(workbook):
    try:
        vba_module = workbook.VBProject.VBComponents.Add(1)  # 1 = vbext_ct_StdModule
        vba_module.CodeModule.AddFromString(VBA_CODE)
    except Exception as e:
        print(f"Error al insertar código VBA: {str(e)}")
        print("Por favor, asegúrese de que 'Confiar en el acceso al modelo de objetos del proyecto de VBA' esté habilitado en la configuración de seguridad de Excel.")
 """

@app.route('/procesar_vba', methods=['GET', 'POST'])
def procesar_vba():
    data=None
    temp_path = None
    if request.method == 'POST':
        file = request.files['file']
        sheet_name = request.form['sheet_name']
        if file and file.filename.endswith('.xlsx'):
            # Guardar el archivo temporalmente
            temp_dir = tempfile.mkdtemp()
            temp_path = os.path.join(temp_dir, 'uploaded_file.xlsx')
            file.save(temp_path)
            # Inicializar COM
            pythoncom.CoInitialize()
            # Ejecutar macros VBA
            excel = None
            try:
                excel = win32com.client.Dispatch("Excel.Application")
                print(f"Excel Application: {excel}")  # Add this line
                workbook = excel.Workbooks.Open(temp_path)

                # Insertar el código VBA en el libro
                insert_vba_code(workbook)

                # Ejecutar las macros
                excel.Application.Run("DesvincularCeldas", sheet_name)
                excel.Application.Run("TabularInformacionCorregido", sheet_name)

                # Guardar y cerrar
                workbook.Save()
                workbook.Close()

                # Esperar un momento para asegurarse de que Excel haya liberado el archivo
                time.sleep(1)

                # Enviar el archivo procesado de vuelta al usuario
                return send_file(temp_path, as_attachment=True, download_name='processed_file.xlsx')
            except Exception as e:
                error_message = str(e)
                if "El acceso mediante programacin al proyecto de Visual Basic no es de confianza" in error_message:
                    error_message += "\n\nPor favor, habilite 'Confiar en el acceso al modelo de objetos del proyecto de VBA' en la configuración de seguridad de Excel."
                return f"Error: {error_message}"
            finally:
                if excel:
                    try:
                        excel.Quit()
                    except Exception as e:
                        print(f"Error al cerrar Excel: {str(e)}")
                pythoncom.CoUninitialize()

                # Intentar eliminar el archivo y el directorio temporal
                for _ in range(5):  # Intentar hasta 5 veces
                    try:
                        if os.path.exists(temp_path):
                            os.remove(temp_path)
                        if os.path.exists(temp_dir):
                            os.rmdir(temp_dir)
                        break
                    except PermissionError:
                        time.sleep(1)  # Esperar un segundo antes de intentar de nuevo
    if temp_path:
        data = pd.read_excel(temp_path)
    return render_template('vba.html', data=data)

@app.route('/procesar_pandas', methods=['POST'])
def procesar_pandas():
    if request.method == 'POST':
        file = request.files['file']
        if file:
            filename = file.filename
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            # Procesar el archivo Excel con pandas
            try:
                # Nombre de la hoja siempre será 'Tabla Tabulada'
                sheet_name = 'Tabla Tabulada'
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                # Eliminar filas donde no hay tiendas registradas o datos esenciales vacíos
                df = df.dropna(subset=['N° de Local', 'Nombre', 'Apertura', 'Cierre'])
                # 1. Eliminar filas duplicadas
                df = df.drop_duplicates(subset=['N° de Local', 'Nombre', 'Apertura', 'Cierre', 'Día Semana 1', 'Fecha 1'])
                # 2. Corregir formato de fechas (y aplicar el formato deseado: dd-mm-yyyy)
                df['Fecha 1'] = pd.to_datetime(df['Fecha 1'], errors='coerce')
                df = df.dropna(subset=['Fecha 1'])  # Eliminar filas con fechas inválidas
                df['Fecha 1'] = df['Fecha 1'].dt.strftime('%d-%m-%Y')  # Formatear la fecha como dd-mm-yyyy
                # Guardar el archivo corregido y ajustar el formato de las celdas de fecha
                output_filename = 'processed_file_corregido.xlsx'
                output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
                
                # Utilizar el writer con xlsxwriter para ajustar el ancho de columnas
                writer = pd.ExcelWriter(output_path, engine='xlsxwriter')
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                # Ajustar el ancho de las columnas
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                for column in df:
                    column_width = max(df[column].astype(str).map(len).max(), len(column))
                    col_idx = df.columns.get_loc(column)
                    worksheet.set_column(col_idx, col_idx, column_width)
                writer.close()  # Usar 'close' en lugar de 'save'
                return send_file(output_path, as_attachment=True)
            except Exception as e:
                return f"Error al procesar el archivo: {e}"
    return render_template('vba.html')

@app.route('/limpiar_datos_adicionales', methods=['POST'])
def limpiar_datos_adicionales():
    if request.method == 'POST':
        file = request.files['file']
        if file:
            filename = file.filename
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)

            try:
                # Cargar el archivo Excel
                df = pd.read_excel(file_path, sheet_name='Tabla Tabulada')

                # Eliminar columnas "Fecha 2" y "Día Semana 2" si existen
                columnas_a_eliminar = ['Fecha 2', 'Día Semana 2']
                df = df.drop(columns=[col for col in columnas_a_eliminar if col in df])

                # Convertir las fechas usando el formato correcto (DD-MM-YYYY)
                for columna in ['Fecha 1']:
                    df[columna] = pd.to_datetime(df[columna], format='%d-%m-%Y', errors='coerce').dt.strftime('%Y-%m-%d')

                # Función para convertir números como 900 a '09:00' y 1000 a '10:00'
                def convertir_a_hora(valor):
                    try:
                        valor_str = f"{int(valor):04d}"
                        return f"{valor_str[:2]}:{valor_str[2:]}"
                    except (ValueError, TypeError):
                        return ''

                # Aplicar la conversión a las columnas "Apertura" y "Cierre"
                for columna in ['Apertura', 'Cierre']:
                    df[columna] = df[columna].apply(convertir_a_hora)

                # Guardar el archivo limpio
                output_filename = 'processed_file_limpio.xlsx'
                output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)

                # Guardar con formato ajustado
                writer = pd.ExcelWriter(output_path, engine='xlsxwriter')
                df.to_excel(writer, sheet_name='Tabla Tabulada', index=False)

                # Ajustar el ancho de las columnas
                workbook = writer.book
                worksheet = writer.sheets['Tabla Tabulada']
                for column in df:
                    column_width = max(df[column].astype(str).map(len).max(), len(column))
                    col_idx = df.columns.get_loc(column)
                    worksheet.set_column(col_idx, col_idx, column_width)

                writer.close()

                return send_file(output_path, as_attachment=True)

            except Exception as e:
                return f"Error al procesar el archivo: {e}"

    return render_template('vba.html')

@app.route('/convertir_xlsx_a_csv', methods=['POST'])
def convertir_xlsx_a_csv():
    if 'file' not in request.files:
        return "No se ha subido ningn archivo", 400

    file = request.files['file']

    if file.filename == '':
        return "No se ha seleccionado ningún archivo", 400

    if file and file.filename.endswith('.xlsx'):
        try:
            # Crear un archivo temporal para guardar el XLSX
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                file.save(tmp.name)
                tmp_path = tmp.name

            # Leer el archivo XLSX
            df = pd.read_excel(tmp_path)

            # Convertir a CSV
            csv_path = tmp_path.replace('.xlsx', '.csv')
            df.to_csv(csv_path, index=False)

            # Preparar el archivo para descarga
            return send_file(csv_path, as_attachment=True, download_name='converted_file.csv', mimetype='text/csv')

        except Exception as e:
            return f"Error al procesar el archivo: {e}", 500

        finally:
            # Eliminar archivos temporales
            if 'tmp_path' in locals():
                try:
                    os.remove(tmp_path)
                except:
                    pass
            if 'csv_path' in locals():
                try:
                    os.remove(csv_path)
                except:
                    pass

    return "Formato de archivo no válido. Por favor, sube un archivo .xlsx", 400

@app.route('/tabular_por_dias', methods=['POST'])
def tabular_por_dias():
    if 'file' not in request.files:
        return 'No file part', 400
    file = request.files['file']
    if file.filename == '':
        return 'No selected file', 400
    
    month = request.form.get('month')
    feriados = request.form.get('feriados', '').split(',')

    # Convertir el mes seleccionado a objeto datetime
    selected_month = datetime.strptime(month, '%Y-%m')
    
    # Leer el archivo
    if file.filename.endswith('.csv'):
        df = pd.read_csv(file)
    elif file.filename.endswith('.xlsx'):
        df = pd.read_excel(file)
    else:
        return 'Formato de archivo no soportado', 400

    # Convertir la columna 'Fecha 1' a datetime
    df['Fecha 1'] = pd.to_datetime(df['Fecha 1'])

    # Filtrar el DataFrame para incluir solo las fechas del mes seleccionado
    df = df[(df['Fecha 1'].dt.year == selected_month.year) & (df['Fecha 1'].dt.month == selected_month.month)]

    # Crear un diccionario para almacenar los datos tabulados
    dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    tabulated_data = {dia: [] for dia in dias_semana + ['Feriado']}

    # Modificar el procesamiento de feriados
    feriados = request.form.get('feriados', '').split(',')
    feriados = [datetime.strptime(f, '%Y-%m-%d').date() if f else None for f in feriados if f]

    # Llenar el diccionario
    for _, row in df.iterrows():
        fecha = row['Fecha 1']
        dia_semana = dias_semana[fecha.weekday()]
        
        # Comparar la fecha completa, no solo el día
        if fecha.date() in feriados:
            tabulated_data['Feriado'].append(row.to_dict())
        else:
            tabulated_data[dia_semana].append(row.to_dict())

    # Crear un DataFrame con los datos tabulados
    result_dfs = []
    for dia, data in tabulated_data.items():
        if data:  # Solo crear un DataFrame si hay datos para ese día
            temp_df = pd.DataFrame(data)
            temp_df['Día'] = dia  # Agregar una columna para identificar el día
            result_dfs.append(temp_df)
    
    if result_dfs:
        result_df = pd.concat(result_dfs, ignore_index=True)
    else:
        result_df = pd.DataFrame(columns=['Día'] + list(df.columns))  # DataFrame vacío con las columnas correctas

    # Crear un buffer en memoria para guardar el archivo CSV
    buffer = io.StringIO()
    result_df.to_csv(buffer, index=False)
    buffer.seek(0)

    # Enviar el archivo como respuesta
    return send_file(
        io.BytesIO(buffer.getvalue().encode()),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'resultado_tabulado_{selected_month.strftime("%Y-%m")}.csv'
    )

@app.route('/obtener_tiendas', methods=['POST'])
def obtener_tiendas():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    try:
        # Si es CSV
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        # Si es Excel
        else:
            df = pd.read_excel(file)
        
        # Obtener lista única de tiendas
        tiendas = df['Nombre'].unique().tolist()
        return jsonify(tiendas)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/configurar_gourmet', methods=['POST'])
def configurar_gourmet():
    if 'file' not in request.files:
        return 'No file uploaded', 400
    
    file = request.files['file']
    tiendas_gourmet = request.form.getlist('tiendas_gourmet')

    try:
        # Obtener el nombre del archivo original
        original_filename = file.filename
        # Extraer la fecha del nombre del archivo si sigue el patrón resultado_tabulado_YYYY-MM.csv
        date_part = original_filename.split('resultado_tabulado_')[-1] if 'resultado_tabulado_' in original_filename else ''
        
        # Leer el archivo
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)

        # Separar registros de feriados
        df_feriados = df[df['Día'] == 'Feriado']
        df_no_feriados = df[df['Día'] != 'Feriado']

        # Para los no feriados, marcar las tiendas gourmet
        df_normal = df_no_feriados[~df_no_feriados['Nombre'].isin(tiendas_gourmet)]
        df_gourmet = df_no_feriados[df_no_feriados['Nombre'].isin(tiendas_gourmet)].copy()
        
        # Cambiar el día a 'Especial' para las tiendas gourmet
        df_gourmet['Día'] = 'Especial'

        # Concatenar en el orden deseado: normal -> feriados -> gourmet
        df_final = pd.concat([df_normal, df_feriados, df_gourmet])

        # Crear un buffer en memoria para el CSV
        output = io.StringIO()
        df_final.to_csv(output, index=False)
        output.seek(0)

        # Crear el nombre del archivo de salida manteniendo la fecha original
        output_filename = f'resultado_tabulado_{date_part}'
        if not output_filename.endswith('.csv'):
            output_filename = f'resultado_tabulado_gourmet.csv'

        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=output_filename
        )

    except Exception as e:
        return str(e), 500

def validate_time_range(apertura, cierre):
    """
    Valida que el rango de tiempo sea coherente.
    
    Args:
        apertura: objeto datetime.time para la apertura
        cierre: objeto datetime.time para el cierre
        
    Returns:
        bool: True si el rango es válido
        str: Mensaje de error si el rango es inválido
    """
    try:
        if not apertura or not cierre:
            return False, "Los horarios no pueden estar vacíos"
            
        # Convertir a minutos para comparación fácil
        apertura_mins = apertura.hour * 60 + apertura.minute
        cierre_mins = cierre.hour * 60 + cierre.minute
        
        # Si el cierre es antes que la apertura, asumimos que cruza medianoche
        if cierre_mins < apertura_mins:
            cierre_mins += 24 * 60
            
        # Validar que la diferencia sea razonable (ej: máximo 24 horas)
        if cierre_mins - apertura_mins > 24 * 60:
            return False, "El horario no puede exceder 24 horas"
            
        return True, ""
        
    except Exception as e:
        return False, f"Error validando horarios: {str(e)}"

#========================================================================================================================================================================================================================================================

""" Proceso de Configracion de datos de Excel  y analisis de datos Sambil la Candelaria ⬇️ """

from flask import Flask, request, render_template, send_file
from openpyxl import load_workbook
import pandas as pd
import matplotlib.pyplot as plt
import os
import plotly.graph_objects as go
import plotly.utils
import json


def sambil_convert_value(val):
    if val is None:
        return ""
    if hasattr(val, 'strftime'):
        return val.strftime('%d-%m-%Y')
    return str(val)

def sambil_guardar_resultado(filepath, sheet_name):
    workbook = load_workbook(filepath)
    
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"La hoja '{sheet_name}' no existe en el archivo.")
    
    sheet = workbook[sheet_name]

    # Desvincular celdas y copiar valores
    for merged_range in list(sheet.merged_cells.ranges):
        main_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
        value = main_cell.value
        sheet.unmerge_cells(str(merged_range))
        
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell.value = value

    # Modificar celdas rojas
    red_rgb = "FFFF0000"
    for row in sheet.iter_rows(min_row=1):  # Comenzar desde la primera fila
        for cell in row:
            if cell.fill and cell.fill.fgColor and hasattr(cell.fill.fgColor, "rgb"):
                if cell.fill.fgColor.rgb == red_rgb:
                    if cell.value:
                        cell.value = f"roja.{cell.value}"
                    else:
                        cell.value = "roja"

    # Crear DataFrame y guardar como CSV
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append([str(cell).strip() if cell is not None else '' for cell in row])
    
    df = pd.DataFrame(data)
    # Usar la primera fila como encabezado
    df.columns = df.iloc[0]
    df = df.iloc[1:]
    
    csv_filepath = filepath.replace('.xlsx', '_resultado.csv')
    df.to_csv(csv_filepath, index=False, encoding='utf-8')

    return csv_filepath

def sambil_analizar_y_graficar(csv_filepath):
    try:
        # Primero intentamos leer con múltiples filas de encabezado
        df = pd.read_csv(csv_filepath, header=[0,1,2,3], on_bad_lines='skip')
    except Exception as e:
        print(f"Error al leer con múltiples encabezados: {e}")
        try:
            # Si falla, intentamos leer de manera más simple
            df = pd.read_csv(csv_filepath, on_bad_lines='skip')
        except Exception as e:
            print(f"Error al leer el CSV: {e}")
            raise ValueError("No se pudo leer el archivo correctamente")

    # Normalizar los nombres de las columnas
    if isinstance(df.columns, pd.MultiIndex):
        # Para columnas multi-índice, combinar niveles no nulos
        df.columns = [
            '_'.join([str(level) for level in col if pd.notna(level) and str(level).strip() != ''])
            for col in df.columns.values
        ]
    
    # Limpiar nombres de columnas
    df.columns = [col.strip().upper().replace(' ', '_') for col in df.columns]
    
    # Convertir todos los valores a string y manejar valores nulos
    df = df.fillna('')
    df = df.applymap(lambda x: str(x).strip())
    
    # Identificar las filas que contienen 'roja'
    row_mask = df.apply(lambda row: row.str.contains('roja', case=False, na=False, regex=False).any(), axis=1)
    
    # Identificar las columnas que contienen 'roja'
    col_mask = df.apply(lambda col: col.str.contains('roja', case=False, na=False, regex=False).any(), axis=0)
    
    # Columnas importantes que queremos mantener
    columnas_base = [
        'NIVEL', 'LOCAL', 'NOMBRE', 'FECHA',
        'PISO', 'AREA', 'SECTOR', 'CATEGORIA',
        'ESTADO', 'OBSERVACION', 'RESPONSABLE', 'PRIORIDAD',
        'TIPO', 'UBICACION', 'CODIGO', 'DESCRIPCION',
        'ESTATUS', 'COMENTARIOS'
    ]
    
    # Buscar columnas importantes de manera más flexible
    columnas_importantes = []
    for col in df.columns:
        if any(base in col.upper().replace('Á', 'A').replace('É', 'E').replace('Í', 'I')
               .replace('Ó', 'O').replace('Ú', 'U') for base in columnas_base):
            columnas_importantes.append(col)
    
    # Si no encontramos columnas importantes, incluir todas las columnas
    if not columnas_importantes:
        columnas_importantes = list(df.columns)
    
    # Crear máscara final para columnas
    col_mask_final = pd.Series(False, index=df.columns)
    col_mask_final[columnas_importantes] = True
    col_mask_final = col_mask_final | col_mask
    
    # Filtrar el DataFrame
    df_filtered = df[row_mask].loc[:, col_mask_final]
    
    # Contar celdas rojas por local
    df['Red_Count'] = df.apply(lambda row: row.str.contains('roja', case=False, na=False, regex=False).sum(), axis=1)
    
    # Encontrar la columna del local de manera más flexible
    local_columns = [col for col in df.columns if 'LOCAL' in col.upper()]
    if not local_columns:
        # Si no encuentra columna LOCAL, usar la primera columna como fallback
        local_column = df.columns[0]
    else:
        local_column = local_columns[0]
    
    # Filtrar solo los locales que tienen celdas rojas
    red_counts = df.groupby(local_column)['Red_Count'].sum()
    red_counts = red_counts[red_counts > 0]
    red_counts = red_counts.sort_values(ascending=False)

    # Crear las gráficas
    n = 14
    grupos_locales = [red_counts[i:i + n] for i in range(0, len(red_counts), n)]
    
    figuras = []
    for i, grupo in enumerate(grupos_locales):
        fig = go.Figure(data=[
            go.Bar(
                x=grupo.index,
                y=grupo.values,
                marker_color='red',
                text=grupo.values,
                textposition='auto',
            )
        ])

        fig.update_layout(
            title=f'Locales con Celdas Rojas (Grupo {i+1})',
            xaxis_title='Locales',
            yaxis_title='Cantidad de Celdas Rojas',
            paper_bgcolor='white',
            plot_bgcolor='white',
            height=400,
            xaxis={
                'tickangle': 45,
                'tickfont': {'size': 12},
            },
            margin=dict(t=50, l=50, r=30, b=100),
            showlegend=False
        )

        figuras.append(json.dumps(fig, cls=plotly.utils.PlotlyJSONEncoder))

    # Generar HTML con estilo para celdas rojas
    def add_red_style(df):
        html = df.to_html(classes='table table-striped', index=False)
        html = html.replace('>roja<', ' class="celda-roja">roja<')
        html = html.replace('>roja.', ' class="celda-roja">roja.')
        return html

    table_html = add_red_style(df_filtered)

    return table_html, figuras

@app.route('/descargar_csv/<filename>')
def sambil_descargar_csv(filename):
    try:
        return send_file(filename,
                        mimetype='text/csv',
                        as_attachment=True,
                        attachment_filename='resultado.csv')
    except Exception as e:
        return str(e)

@app.route('/sambil')
def sambil_index():
    return render_template('sambil_index.html')

@app.route('/procesar_sambil', methods=['POST'])
def sambil_procesar():
    file = request.files['archivo']
    sheet_name = request.form['sheet_name']
    filepath = f"temp_{file.filename}"
    file.save(filepath)

    try:
        csv_filepath = sambil_guardar_resultado(filepath, sheet_name)
        table_html, chart_data = sambil_analizar_y_graficar(csv_filepath)
        return render_template('sambil_resultado.html', 
                            table_html=table_html, 
                            chart_data=chart_data,
                            csv_filename=csv_filepath)
    except ValueError as e:
        if os.path.exists(filepath):
            os.remove(filepath)
        return str(e), 400


""" otros """

def diagnostico_excel_server():
    """Función para diagnosticar problemas con Excel en el servidor"""
    messages = []
    try:
        pythoncom.CoInitialize()
        messages.append("✓ CoInitialize exitoso")
        
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            messages.append(f"✓ Excel iniciado correctamente - Versión: {excel.Version}")
        except Exception as e:
            messages.append(f"✗ Error al iniciar Excel: {str(e)}")
            raise
            
        try:
            excel.Visible = False
            excel.DisplayAlerts = False
            messages.append("✓ Configuración de Excel establecida")
        except Exception as e:
            messages.append(f"✗ Error al configurar Excel: {str(e)}")
            raise
            
        # Verificar permisos y rutas
        temp_dir = tempfile.gettempdir()
        messages.append(f"Directorio temporal: {temp_dir}")
        messages.append(f"Permisos de escritura en temp: {os.access(temp_dir, os.W_OK)}")
        
        # Verificar usuario que ejecuta el proceso
        messages.append(f"Usuario actual: {os.getenv('USERNAME')}")
        
        # Verificar registro de Excel
        try:
            import winreg
            key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Office\16.0\Excel\InstallRoot")
            path = winreg.QueryValueEx(key, "Path")[0]
            messages.append(f"✓ Ruta de Excel en registro: {path}")
        except Exception as e:
            messages.append(f"✗ Error al verificar registro de Excel: {str(e)}")
        
        return True, messages
    except Exception as e:
        messages.append(f"Error general: {str(e)}")
        return False, messages
    finally:
        try:
            if 'excel' in locals():
                excel.Quit()
        except:
            pass
        pythoncom.CoUninitialize()

@app.route('/diagnostico_excel')
def diagnostico():
    success, messages = diagnostico_excel_server()
    return render_template('diagnostico.html', 
                         success=success, 
                         messages=messages)

if __name__ == '__main__':
    app.run(host='192.168.102.17', port=5000, debug=True)